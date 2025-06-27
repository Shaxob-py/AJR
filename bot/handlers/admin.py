import re
from datetime import datetime
from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, ReplyKeyboardRemove, FSInputFile
from openpyxl import Workbook
from sqlalchemy import select, update
from pathlib import Path
from db.models import Admin, Customer, async_session_maker, Payment
from bot.despecher.state import AdminState
from bot.keys.reply import admin_menu_button,  choice_data_button, reply_button
from bot.teksts.teksts import wrong_password, welcome_admin, choice_data, data, \
    back, main_menu, all_users, all_daily, all_payments, change_password, warning_change_password, changing_password, \
    return_password

admin_router = Router()

@admin_router.message(F.text == "admin")
async def admin_panel(message: Message, state: FSMContext):
    async with async_session_maker() as session:
        admin = await session.get(Admin, message.from_user.id)
    if admin:
        await message.answer('Parolni kriting')
        await state.set_state(AdminState.register_admin)
        await message.delete()


@admin_router.message(AdminState.register_admin, F.text)
async def show_stats(message: Message, state: FSMContext):
    password = message.text
    async with async_session_maker() as session:
        admin = await session.get(Admin, message.from_user.id)
        if admin.password != password:
            await message.answer(wrong_password)
            await message.delete()
            await state.clear()
            return
        await message.delete()
        await state.set_state(AdminState.menu_panel)
        await message.answer(welcome_admin, reply_markup=admin_menu_button())


@admin_router.message(AdminState.data_state, F.text == back)
@admin_router.message(AdminState.change_password_state, F.text == back)
async def show_menu(message: Message, state: FSMContext):
    await message.delete()
    await message.answer(main_menu, reply_markup=admin_menu_button())
    await state.set_state(AdminState.menu_panel)


@admin_router.message(AdminState.menu_panel, F.text == data)
async def send_broadcast(message: Message, state: FSMContext):
    await message.answer(choice_data, reply_markup=choice_data_button())
    await state.set_state(AdminState.data_state)


@admin_router.message(AdminState.data_state, F.text.in_([all_users, all_daily, all_payments]))
async def send_data(message: Message):
    if message.text == all_users:
        async with async_session_maker() as session:
            result = await session.execute(select(Customer))
            users = result.scalars().all()

            if not users:
                return await message.answer("⛔ Ma'lumot topilmadi.")

            today = datetime.now()
            date_str = today.strftime("%Y-%m-%d")
            safe_date = re.sub(r'[\\/*?:"<>|]', "_", date_str)

            temp_dir = Path("temp_excel")
            temp_dir.mkdir(exist_ok=True)

            filename = f"customers_{safe_date}.xlsx"
            file_path = temp_dir / filename

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Users"

            # Excel header
            sheet.append(["№", "Full Name", "Username", "Name", "Phone", "Location"])

            for i, user in enumerate(users, 1):
                sheet.append([
                    i,
                    user.full_name or "",
                    f"@{user.username}" if user.username else "",
                    user.name or "",
                    user.phone_number or "",
                    user.location or ""
                ])

            workbook.save(file_path)

            await message.answer_document(
                document=FSInputFile(file_path),
                caption=f"Barcha foydalanuvchilar"
            )

            file_path.unlink(missing_ok=True)

    elif message.text == all_payments:
        async with async_session_maker() as session:
            result = await session.execute(select(Payment))
            payments = result.scalars().all()

            if not payments:
                return await message.answer("📭 Hozircha to‘lovlar yo‘q.")

            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"payments_{date_str}.xlsx"
            file_path = Path("temp_excel") / filename
            file_path.parent.mkdir(exist_ok=True)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Payments"

            # Header
            sheet.append([
                "№", "Pay (sum)", "User ID", "Paid Time", "Status",
                "Location", "Coordinates", "Phone Number"
            ])

            for i, payment in enumerate(payments, 1):
                sheet.append([
                    i,
                    float(payment.pay),
                    payment.user_id,
                    payment.paid.strftime("%Y-%m-%d %H:%M"),
                    payment.status or "",
                    payment.location or "",
                    payment.coordinates or "",
                    payment.phone_number or ""
                ])

            workbook.save(file_path)

            await message.answer_document(
                FSInputFile(file_path),
                caption=f"💳 To‘lovlar Excel fayli: {date_str}"
            )

            file_path.unlink(missing_ok=True)

    elif message.text == all_daily:
        async with async_session_maker() as session:
            result = await session.execute(select(Payment))
            payments = result.scalars().all()

            if not payments:
                return await message.answer("📭 Hozircha to‘lovlar yo‘q.")

            date_str = datetime.now().strftime("%Y-%m-%d")
            filename = f"payments_{date_str}.xlsx"
            file_path = Path("temp_excel") / filename
            file_path.parent.mkdir(exist_ok=True)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Payments"

            # Header
            sheet.append([
                "№", "Pay (sum)", "User ID", "Paid Time", "Status",
                "Location", "Coordinates", "Phone Number"
            ])

            for i, payment in enumerate(payments, 1):
                sheet.append([
                    i,
                    float(payment.pay),
                    payment.user_id,
                    payment.paid.strftime("%Y-%m-%d %H:%M"),
                    payment.status or "",
                    payment.location or "",
                    payment.coordinates or "",
                    payment.phone_number or ""
                ])

            workbook.save(file_path)

            await message.answer_document(
                FSInputFile(file_path),
                caption=f"💳 To‘lovlar Excel fayli: {date_str}"
            )

            file_path.unlink(missing_ok=True)

    return None

@admin_router.message(AdminState.menu_panel, F.text==change_password)
async def change_password_admin(message: Message,state: FSMContext):
    await message.answer(warning_change_password)
    await message.answer(changing_password,reply_markup=reply_button([back]))
    await state.set_state(AdminState.change_password_state)

@admin_router.message(AdminState.change_password_state,F.text)
async def change_password_state(message: Message,state: FSMContext):
    await state.update_data(password=message.text)
    await message.answer(return_password,reply_markup=ReplyKeyboardRemove())
    await state.set_state(AdminState.confirm_password_state)

@admin_router.message(AdminState.confirm_password_state, F.text)
async def confirm_password_state(message: Message, state: FSMContext):
    data = await state.get_data()
    password_from_state = data.get("password")

    if not password_from_state or message.text != password_from_state:
        await message.answer("❌ Parol noto‘g‘ri. Iltimos, qaytadan urinib ko‘ring.", reply_markup=admin_menu_button())
        await state.clear()
        return

    async with async_session_maker() as session:
        await session.execute(update(Admin).where(Admin.id == message.from_user.id).values(password=message.text))
        await session.commit()

    await message.answer("✅ Parol tasdiqlandi!", reply_markup=admin_menu_button())
    await state.clear()